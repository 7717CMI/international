"""
Convert the U.S. Banking Industry & International Moving Market Excel dataset
to value.json and volume.json for the dashboard.

Reads from the Master Sheet which has flat structured data:
  Column A: Unit (Value/Volume)
  Column B: Segment type
  Column C: Sub-segment (parent for hierarchical)
  Column D: Sub-segment 1 (leaf for hierarchical)
  Columns E-M: Years 2025-2033

RULE: No double counting - for "By Move Type", only leaf children have data.
Parent nodes (Employee Mobility, Asset Mobility) are structural containers only.
"""
import json
import openpyxl
import os

EXCEL_FILE = "Dataset-U.S. Banking Industry & International Moving Market.xlsx"
YEARS = [str(y) for y in range(2025, 2034)]  # 2025-2033


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Master Sheet"]

    value_data = {"U.S.": {}}
    volume_data = {"U.S.": {}}

    # Read Master Sheet data rows (row 7 onward, header is row 6)
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        unit = str(row[0]).strip() if row[0] else None
        segment_type = str(row[1]).strip() if row[1] else None
        sub_segment = str(row[2]).strip() if row[2] else None
        sub_segment_1 = str(row[3]).strip() if row[3] else None

        if not unit or not segment_type or not sub_segment:
            continue

        # Extract year values (columns E-M = indices 4-12)
        year_values = {}
        for i, year in enumerate(YEARS):
            val = row[4 + i]
            if val is not None:
                if unit == "Value":
                    year_values[year] = round(float(val), 1)
                else:  # Volume
                    year_values[year] = int(round(float(val)))

        if not year_values:
            continue

        # Select target dict
        target = value_data if unit == "Value" else volume_data

        # Handle "By Move Type" hierarchically
        # sub_segment = parent (Employee Mobility / Asset Mobility)
        # sub_segment_1 = leaf child
        if segment_type == "By Move Type":
            if segment_type not in target["U.S."]:
                target["U.S."][segment_type] = {}
            if sub_segment not in target["U.S."][segment_type]:
                target["U.S."][segment_type][sub_segment] = {}
            target["U.S."][segment_type][sub_segment][sub_segment_1] = year_values
        else:
            # Flat segments: use sub_segment as the item name
            if segment_type not in target["U.S."]:
                target["U.S."][segment_type] = {}
            target["U.S."][segment_type][sub_segment] = year_values

    wb.close()

    # Write output
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public", "data")
    with open(os.path.join(out_dir, "value.json"), "w", encoding="utf-8") as f:
        json.dump(value_data, f, indent=2, ensure_ascii=False)
    with open(os.path.join(out_dir, "volume.json"), "w", encoding="utf-8") as f:
        json.dump(volume_data, f, indent=2, ensure_ascii=False)

    # Print summary
    print("=== VALUE DATA ===")
    for seg_type, segments in value_data["U.S."].items():
        if seg_type == "By Move Type":
            print(f"\n{seg_type}:")
            for parent, children in segments.items():
                parent_total = sum(list(children.values())[0].values()) if children else 0
                print(f"  {parent} ({len(children)} children):")
                for child, vals in children.items():
                    print(f"    {child}: {vals.get('2025', '?')} -> {vals.get('2033', '?')}")
        else:
            print(f"\n{seg_type} ({len(segments)} items):")
            for seg, vals in segments.items():
                print(f"  {seg}: {vals.get('2025', '?')} -> {vals.get('2033', '?')}")

    print("\n=== VOLUME DATA ===")
    for seg_type, segments in volume_data["U.S."].items():
        if seg_type == "By Move Type":
            print(f"\n{seg_type}:")
            for parent, children in segments.items():
                print(f"  {parent} ({len(children)} children):")
                for child, vals in children.items():
                    print(f"    {child}: {vals.get('2025', '?')} -> {vals.get('2033', '?')}")
        else:
            print(f"\n{seg_type} ({len(segments)} items):")
            for seg, vals in segments.items():
                print(f"  {seg}: {vals.get('2025', '?')} -> {vals.get('2033', '?')}")

    # Double-count verification
    print("\n=== DOUBLE-COUNT CHECK (Value 2025) ===")
    for seg_type in value_data["U.S."]:
        total = 0
        if seg_type == "By Move Type":
            for parent, children in value_data["U.S."][seg_type].items():
                for child, vals in children.items():
                    total += vals.get("2025", 0)
        else:
            for seg, vals in value_data["U.S."][seg_type].items():
                total += vals.get("2025", 0)
        print(f"  {seg_type}: {total:.1f}")

    print("\nDone! Files written to public/data/")


if __name__ == "__main__":
    main()
